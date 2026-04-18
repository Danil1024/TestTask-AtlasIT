from io import BytesIO

from django.db.models import QuerySet, Sum, Count
from django.db.models.functions import TruncDate

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from reports.models import Order, Customer


def generate_revenue_report() -> BytesIO:
    queryset = (
        Order.objects
        .filter(status=Order.Status.PAID)
        .annotate(date=TruncDate('created_at'))
        .values('date')
        .annotate(
            total_revenue=Sum('amount', default=0),
            orders_count=Count('id'),
        )
        .order_by('date')
    )

    return _build_xlsx(
        queryset,
        sheet_title='Revenue',
        headers=['date', 'total_revenue', 'orders_count'],
        fields=['date', 'total_revenue', 'orders_count'],
    )


def generate_customers_report() -> BytesIO:
    queryset = (
        Customer.objects
        .annotate(
            total_orders=Count('orders'),
            total_spent=Sum('orders__amount', default=0),
        )
        .values('id', 'name', 'total_orders', 'total_spent')
        .order_by('id')
    )

    return _build_xlsx(
        queryset,
        sheet_title='Customers',
        headers=['customer_id', 'name', 'total_orders', 'total_spent'],
        fields=['id', 'name', 'total_orders', 'total_spent'],
    )


def _build_xlsx(
    queryset: QuerySet,
    sheet_title: str,
    headers: list[str],
    fields: list[str],
) -> BytesIO:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet_title)

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    ws.append(headers)

    for row in queryset.iterator(chunk_size=1000):
        ws.append([row[f] for f in fields])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer